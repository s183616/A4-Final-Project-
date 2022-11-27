# -*- coding: utf-8 -*-
"""
Created on Thu Nov 10 10:07:23 2022

@author: DKISVA
"""
###############################################################################
############################File path directory################################
###############################################################################
import os 
dir_path = os.path.dirname(os.path.realpath(__file__))

###############################################################################
############################ IFC data extraction ##############################
###############################################################################
import ifcopenshell
file = ifcopenshell.open(os.path.join(dir_path,r'model\Duplex.ifc'))

#Define beams
beams = file.by_type("IfcBeam")

for beam in beams:
  lengthbeams=len(beam)


class_beam = [] 
object_type_beam = []
material_beam = []

for beam in beams:
    #Defining IfcClass as 'Beam':
    class_beam.append('Beam')
    
    #Finding all beam object types:
    object_type_beam.append(str(beam.ObjectType)) 

    #Find beam materials:
    for relAssociatesMaterials in beam.HasAssociations:
        material_beam.append(relAssociatesMaterials.RelatingMaterial.Name)

#Reducing of list
res = {}

for i in object_type_beam:
    res[i] = object_type_beam.count(i) 

#List of all beam types without repetitions:
beam_type = list(res.keys())
#Count beam occurances:
beam_count = list(res.values())

#resize class list:
class_beam = class_beam[0:len(beam_type)]
#resize material list:
material_beam = material_beam[0:len(beam_type)]

###############################################################################
############################# IFC data matrix #################################
###############################################################################
#Making a matrix with all information retrieved from IFC-file:

import numpy as np

A = np.column_stack((class_beam,beam_type,beam_count,material_beam))

###############################################################################
############################### Import CSV file ###############################
###############################################################################

import pandas as pd
 
# open the file in read mode
filename1 = pd.read_csv(os.path.join(dir_path,r'input\gen_dk\element_categories.csv'), sep=";")
filename2=filename1.values.tolist()

#print(filename[0][:51])
for row in filename2:
    if row[0]==str(A[0][0]):
        element_cat=(row[1])

###############################################################################
############################### Lifespan calc #################################
###############################################################################

import openpyxl

# open the file in read mode
Data = openpyxl.load_workbook(os.path.join(dir_path,r'input\gen_dk\BUILD_levetidstabel_november_2022.xlsx'))
Data_sheet = Data["Bygningsdele SfB"]
x = "Bærende konstruktioner, øvrige (søjler, bjælker, rammer, skaller)"
#s = x.__contains__("bjælke")
for i in Data_sheet.iter_rows():
    id = i[4].value
    #print(i[4].value)
    if i[4].value == x:
        row1 = i[4].row
        #print(row1)
for j in Data_sheet.iter_cols():
    id = j[1].value
    #print(j[1].value)
    if j[1].value == "Jern, stål og& rustfrit stål":
        col1=j[6].column
        #print(col1)
########################## Locating lifespan ##################################
lifespan = Data_sheet.cell(row1, col1).value
#print(lifespan)


###############################################################################
########################### Link between IFC and CSV ##########################
###############################################################################

import pandas as pd
 
# open the file in read mode
filename3 = pd.read_csv(os.path.join(dir_path,r'input\gen_dk\stages.csv'), sep=";")

filename4=filename3.values.tolist()

for row in filename4:
    if row[0]==element_cat:
        stage_id=(row[1])


###############################################################################
############################## Make unique ID #################################
###############################################################################

import uuid

###############################################################################
############################## Make json folder ###############################
###############################################################################

import os.path

#SET PATH FOR LOCATION OF FOLDER:
#parent_dir = r"C:\Users\DKISVA\OneDrive - Sweco AB\Privat\DTU\Advanced BIM\FINAL HANDIN"
parent_dir = dir_path
#GENERATE FOLDER "JSON FILES":
directory = "output"
final_directory = os.path.join(parent_dir, r'output')
if not os.path.exists(final_directory):
   os.makedirs(final_directory)
print("Directory '% s' is created" % directory)

folder_dir = "%s/%s" % (parent_dir, directory)

#CHANGE WORKING DIRECTORY TO GENERATED FOLDER:
os.chdir(folder_dir)
print("Working directory has been changed to % s" % folder_dir)


###############################################################################
########################## Make ID for all nodes ##############################
###############################################################################

ID_El=str(uuid.uuid1())
ID_Con=str(uuid.uuid1())
ID_Prod=str(uuid.uuid1())
ID_Stage=str(uuid.uuid1())

###############################################################################
################## Genereation of json file_CategoryToElement #################
###############################################################################
ID_CatToEl=str(uuid.uuid1())

CatToEltrue=True

import json

with open('CategoryToElement.json', 'w') as f:
    print("The json file is created")

dictionary = [
 {
    "Edge": [
      {
        "CategoryToElement": {
          "id": ID_CatToEl,
		  "excluded_scenarios": []
        }
      },
      element_cat,
      ID_El
    ]
  }
]
########################### File in folder ####################################
# Serializing json
json_object = json.dumps(dictionary, indent=4)
 
# Writing to sample.json
with open("CategoryToElement.json", "w") as outfile:
    outfile.write(json_object)
    
###############################################################################
###################### Genereation of json file_Element #######################
###############################################################################

Eltrue=True

import json

with open('Element.json', 'w') as f:
    print("The json file is created")

dictionary = [
  {
    "Node": {
      "Element": {
        "id": ID_El,
        "name": {
          "Danish": "Bjælker",
          "English": str(A[0][0]),
          "German": ""
        },
        "source": "User",
        "comment": "",
        "enabled": Eltrue,
        "excluded_scenarios": []
      }
    }
  }
]

########################### File in folder ####################################
# Serializing json
json_object = json.dumps(dictionary, indent=4)
 
# Writing to sample.json
with open("Element.json", "w") as outfile:
    outfile.write(json_object)
    
###############################################################################
############### Genereation of json file_Construction #########################
###############################################################################

Contrue=True

import json

with open('Construction.json', 'w') as f:
    print("The json file is created")


dictionary =   [
  {
    "Node": {
      "Construction": {
        "id": ID_Con,
        "name": {
          "Danish": str(A[0][1]),
          "English": str(A[0][1])
        },
        "unit": "M",
        "source": "User",
        "comment": "Test",
		"locked": Contrue
      }
    }
  }
]

########################### File in folder ####################################
# Serializing json
json_object = json.dumps(dictionary, indent=4)
 
# Writing to sample.json
with open("Construction.json", "w") as outfile:
    outfile.write(json_object)    
########################### ADD more constructions to file ####################
if len(beam_type)>1:

    with open('Construction.json') as fp:
        listObj = json.load(fp)
 

    for t in range(1,len(beam_type)):
        ID_Con_t=str(uuid.uuid1())
        listObj.append({
            "Node": {
              "Construction": {
                "id": ID_Con_t,
                "name": {
                  "Danish": str(A[t][1]),
                  "English": str(A[t][1])
                },
                "unit": "M",
                "source": "User",
                "comment": "Test",
        		"locked": Contrue
              }
            }
          })
#print(listObj)
 
with open('Construction.json', 'w') as json_file:
    json.dump(listObj, json_file, 
                        indent=4,  
                        separators=(',',': '))
 
#print('Successfully appended to the JSON file')

###############################################################################
############### Genereation of json file_ElementToConstruction ################
###############################################################################

ID_ElToCon=str(uuid.uuid1())
ElToContrue=True

import json

with open('ElementToConstruction.json', 'w') as f:
    print("The json file is created")

dictionary = [
  {
    "Edge": [
      {
        "ElementToConstruction": {
          "id": ID_ElToCon,
          "amount": lengthbeams,
          "enabled": ElToContrue,
          "excluded_scenarios": []
        }
      },
      ID_El,
      ID_Con
    ]
  }
]

########################### File in folder ####################################
# Serializing json
json_object = json.dumps(dictionary, indent=4)
 
# Writing to sample.json
with open("ElementToConstruction.json", "w") as outfile:
    outfile.write(json_object)
########################### ADD more elementtoconstruction edges to file ######
if len(beam_type)>1:

    with open('ElementToConstruction.json') as fp:
        listObj = json.load(fp)
 

    for t in range(1,len(beam_type)):
        ID_ElToCon_t=str(uuid.uuid1())
        listObj.append({
          "Edge": [
            {
              "ElementToConstruction": {
                "id": ID_ElToCon_t,
                "amount": lengthbeams,
                "enabled": ElToContrue,
                "excluded_scenarios": []
              }
            },
            ID_El,
            ID_Con_t
          ]
        })
#print(listObj)
 
with open('ElementToConstruction.json', 'w') as json_file:
    json.dump(listObj, json_file, 
                        indent=4,  
                        separators=(',',': '))
 
#print('Successfully appended to the JSON file')
     
###############################################################################
##################### Genereation of json file_Product ########################
###############################################################################

#print(filename4[0][:51])  
#for row in filename4:
#    if "StÃ¥lprofil" == element_cat: 
 
# dropping null value columns to avoid errors
ProdToStagtrue=True
ID_ProdtoStag=str(uuid.uuid1())
ID_ProdtoStag1=str(uuid.uuid1())

filename3 = pd.read_csv(os.path.join(dir_path,r'input\gen_dk\stages.csv'), sep=";")

materials = file.by_type("IfcMaterial")

for material in materials:
    if "steel" in material.Name.lower():
        materialifc=filename3.loc[filename3.name.str.contains('stålprofil',case=False)] 
        Material_ID_A = materialifc["id"].iat[0]
        Material_ID_C = materialifc["id"].iat[1]
        

import json

with open('Product.json', 'w') as f:
    print("The json file is created")

dictionary =   [
  {
    "Node": {
      "Product": {
        "id": ID_Prod,
        "name": {
          "English": str(A[0][3]),
          "German": "",
          "Danish": str(A[0][3])
        },
        "source": "User",
        "comment": "",
        "uncertainty_factor": 1.0,
        "uncertainty_factor_dgnb": 1.3
      }
    }
  } ,
    {
        "Edge": [
            {
                "ProductToStage": {
                    "id": ID_ProdtoStag,
                    "excluded_scenarios": [],
                    "enabled": ProdToStagtrue
                 }
            } ,
            ID_Prod,
            Material_ID_A
        ]
  },
    {
        "Edge": [
            {
                 "ProductToStage": {
                     "id": ID_ProdtoStag1,
                     "excluded_scenarios": [],
                     "enabled": ProdToStagtrue
                  }
            } ,
            ID_Prod,
            Material_ID_C
        ]
	  }
]
########################### File in folder ####################################
# Serializing json
json_object = json.dumps(dictionary, indent=4)
 
# Writing to sample.json
with open("Product.json", "w") as outfile:
    outfile.write(json_object)
    
###############################################################################
############### Genereation of json file_ConstructionToProduct ################
###############################################################################

ID_ConToProd=str(uuid.uuid1())
ConToProdtrue=True
ConToProdfalse=False
import json

with open('ConstructionToProduct.json', 'w') as f:
    print("The json file is created")

dictionary =   [
	{
		"Edge": [
		  {
			"ConstructionToProduct": {
			  "id": ID_ConToProd,
			  "amount": 0.44,
			  "unit": "M3",
			  "lifespan": lifespan,
			  "demolition": ConToProdfalse,
			  "delayed_start": 0,
			  "enabled": ConToProdtrue,
			  "excluded_scenarios": []
			}
		  },
		  ID_Con,
		  ID_Prod
		]
	  }
]

########################### File in folder ####################################
# Serializing json
json_object = json.dumps(dictionary, indent=4)
 
# Writing to sample.json
with open("ConstructionToProduct.json", "w") as outfile:
    outfile.write(json_object)    

########################### ADD more ConstructionToProduct edges to file ######
if len(beam_type)>1:

    with open('ConstructionToProduct.json') as fp:
        listObj = json.load(fp)
 

    for t in range(1,len(beam_type)):
        ID_ConToProd_t=str(uuid.uuid1())
        listObj.append({
    		"Edge": [
    		  {
    			"ConstructionToProduct": {
    			  "id": ID_ConToProd_t,
    			  "amount": 0.44,
    			  "unit": "M3",
    			  "lifespan": lifespan,
    			  "demolition": ConToProdfalse,
    			  "delayed_start": 0,
    			  "enabled": ConToProdtrue,
    			  "excluded_scenarios": []
    			}
    		  },
    		  ID_Con_t,
    		  ID_Prod
    		]
    	  })
#print(listObj)
 
with open('ConstructionToProduct.json', 'w') as json_file:
    json.dump(listObj, json_file, 
                        indent=4,  
                        separators=(',',': '))
 
#print('Successfully appended to the JSON file')
    
###############################################################################
############### Genereation of json file_Building #############################
###############################################################################

ID_Build=str(uuid.uuid1())
ID_BuildRoot=str(uuid.uuid1())
ID_Root=str(uuid.uuid1())

ElToContrue=True

import json

with open('Building.json', 'w') as f:
    print("The json file is created")

dictionary = [
  {
    "Node": {
      "Building": {
        "id": ID_Build,
        "scenario_name": "Original bygningsmodel",
        "locked": "Unlocked",
        "description": "",
        "building_type": "Office",
        "heated_floor_area": 10.0,
        "gross_area": 10.0,
        "gross_area_above_ground": 0.0,
        "storeys_above_ground": 0,
        "storeys_below_ground": 0,
        "storey_height": 0.0,
        "initial_year": 2020,
        "calculation_timespan": 50,
        "calculation_mode": "SC",
        "outside_area": 0.0,
        "plot_area": 0.0,
        "energy_class": "LowEnergy",
        "name": {
          "Danish": ""
        },
        "address": "",
        "owner": "",
        "lca_advisor": "",
        "building_regulation_version": ""
      }
    }
  }
]

########################### File in folder ####################################
# Serializing json
json_object = json.dumps(dictionary, indent=4)
 
# Writing to sample.json
with open("Building.json", "w") as outfile:
    outfile.write(json_object)
    
